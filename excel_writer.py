import xlsxwriter # writing to excel
from openpyxl import load_workbook
import os
from datetime import datetime
import pandas



########################

# write to excel file

def write_to_excel(folder_name, filename, sheetname, experiment_name, heading: list, content: list):
    
    # select file path
    os.makedirs(folder_name, exist_ok=True)
    file_path = os.path.join(folder_name, filename)

    # create sheet
    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet(sheetname)
    
    # write experiment name
    worksheet.write(0,0, experiment_name)

    
    # write heading (list), give information on variables, etc
    row_start = 2

    for row in range(len(heading)):
            worksheet.write(row_start + row, 0, heading[row])

    # write content (content list of iterables)
    row_start += (2 + len(heading))
    column_start = 0

    for row in range(len(content)):
        for column in range(len(content[row])):
            worksheet.write(row_start + row, column_start + column, content[row][column])

    # close workbook
    workbook.close()



def annotate_excel_file_results(folder_name, file_name, scenario, factor, error, table, query_number, run, total_runs, time_original, time_distorted, tuple_amount_original, tuple_amount_distorted, tuple_amount_intersection):
    
    # current date and time
    today = datetime.now().strftime("%Y_%m_%d")
    current = datetime.now().strftime("%H_%M_%S")

    file_path = f"{folder_name}\{file_name}"
    
    # Load the existing workbook
    workbook = load_workbook(file_path)
    sheet = workbook["results"]

    if tuple_amount_original == 0:
         recall = "X"
    else:
         recall = tuple_amount_intersection/tuple_amount_original

    if tuple_amount_distorted == 0:
         precission = 0
    else:
         precission = tuple_amount_intersection/tuple_amount_distorted

    # Append a new row
    sheet.append([today, current, scenario, factor, error, table, f"Q{query_number}", run, total_runs, time_original, time_distorted, tuple_amount_original, tuple_amount_distorted, tuple_amount_intersection, recall, precission])

    # Save the changes
    workbook.save(file_path)





def write_excel_file_query_results(scenario, folder_name, query_number, query, runs, factor, error, accuracies, accuracies_stats, all_accuracies, precissions, precissions_stats, all_precissions, average_query_original_output_amounts, average_query_distorted_output_amounts, average_original_times, average_distorted_times, tables):

    # current date and time
    today = datetime.now().strftime("%Y_%m_%d")
    current = datetime.now().strftime("%H_%M_%S")

    # Experiment results

    filename = "Data_quality_effects_results_" + "query_" + str(query_number) + "___scale_" + str(factor) + "__error_" + str(error) + "_" + str(today) + "---" + str(current) + ".xlsx"
    sheetname = "Experiment"
    experiment_name = f"TPC-H benchmark queries with scaling factor {factor} and {scenario} error {error}%"


    experiment_details = []

    # create content for Excel
    content = [["Benchmark query number " + str(query_number) + " : " + query ], [], [f"Number of runs: {runs}"], []]
    content.append(["Table:", ""] + tables)
    content.append([])
    content.append(["Average time for original query (ms): ", ""] + average_original_times)
    content.append(["Average time for distorted query (ms): ", ""] + average_distorted_times)
    content.append([])
    content.append(["Average original output amount: ", ""] + average_query_original_output_amounts)
    content.append(["Average distorted output amount: ", ""] + average_query_distorted_output_amounts)
    content.append([])    
    content.append(["Recall:", ""] + accuracies)
    content.append(["Precission:", ""] + precissions)
    content.append([])
    content.append(["Recall stats:"])
    content.append(["Min:", ""] + accuracies_stats[0])
    content.append(["Mean:", ""] + accuracies_stats[1])
    content.append(["Median:", ""] + accuracies_stats[2])
    content.append(["Max:", ""] + accuracies_stats[3])
    content.append([])
    content.append(["Precission stats:"])
    content.append(["Min:", ""] + precissions_stats[0])
    content.append(["Mean:", ""] + precissions_stats[1])
    content.append(["Median:", ""] + precissions_stats[2])
    content.append(["Max:", ""] + precissions_stats[3])
    content.append([])
    content.append([])
    content.append([])
    content.append(["Raw Recall data:"])
    for index, table in enumerate(tables):
         content.append([table] + all_accuracies[index])
    content.append([])
    content.append(["Raw Precission data:"])
    for index, table in enumerate(tables):
         content.append([table] + all_precissions[index])


    # writing to file
    write_to_excel(folder_name, filename, sheetname, experiment_name, experiment_details, content)



def write_overview_from_dataframe_accuracies(folder_name, dataframe, factor, error):


     
    # current date and time
    today = datetime.now().strftime("%Y_%m_%d")
    current = datetime.now().strftime("%H_%M_%S")

    filename = "Data_quality_effects_accuracy_overview_results_" + "__scale_" + str(factor) + "__error_" + str(error) + "_" + str(today) + "---" + str(current) + ".xlsx"

    # select file path
    os.makedirs(folder_name, exist_ok=True)
    file_path = os.path.join(folder_name, filename)

    dataframe.to_excel(file_path)


def write_overview_from_dataframe_times(folder_name, original_time_dataframe, distorted_time_dataframe, factor, error):


     
    # current date and time
    today = datetime.now().strftime("%Y_%m_%d")
    current = datetime.now().strftime("%H_%M_%S")

    filename_originals = "Data_quality_effects_original_times_overview_results_" + "__scale_" + str(factor) + "__error_" + str(error) + "_" + str(today) + "---" + str(current) + ".xlsx"

    filename_distorted = "Data_quality_effects_distorted_times_overview_results_" + "__scale_" + str(factor) + "__error_" + str(error) + "_" + str(today) + "---" + str(current) + ".xlsx"

    # select file path and write to excel
    os.makedirs(folder_name + "/original", exist_ok=True)
    file_path = os.path.join(folder_name + "/original", filename_originals)
    original_time_dataframe.to_excel(file_path, index=False)

    os.makedirs(folder_name + "/distorted", exist_ok=True)
    file_path = os.path.join(folder_name + "/distorted", filename_distorted)
    distorted_time_dataframe.to_excel(file_path, index=False)